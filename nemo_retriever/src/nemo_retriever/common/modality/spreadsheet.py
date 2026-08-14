# SPDX-License-Identifier: Apache-2.0

"""Native spreadsheet extraction for workbook and delimited-text inputs.

The spreadsheet path intentionally does not rasterize a workbook as its first
step.  Cell values, formulas, sheet names and ranges are more reliable
retrieval provenance than OCR over a rendered spreadsheet.  Rendering/OCR of
embedded images remains an explicit follow-up stage rather than a second copy
of the native cell text.
"""

from __future__ import annotations

import csv
import hashlib
import io
import math
import shutil
import subprocess
import tempfile
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Sequence

import pandas as pd

from nemo_retriever.common.params import TextChunkParams


SPREADSHEET_EXTENSIONS = frozenset({".xlsx", ".xls", ".csv"})
WORKBOOK_EXTENSIONS = frozenset({".xlsx", ".xls"})


def _safe_scalar(value: Any) -> str:
    """Turn a workbook/CSV value into stable, embedding-safe text."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _column_name(index: int) -> str:
    """Return an Excel-style column name for a one-based column index."""
    result = ""
    value = int(index)
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def _column_number(label: str) -> int:
    value = 0
    for character in label.upper():
        value = value * 26 + ord(character) - 64
    return value


def _cell_range(row_start: int, row_end: int, col_start: int, col_end: int) -> str:
    return f"{_column_name(col_start)}{row_start}:{_column_name(col_end)}{row_end}"


def _formula_cells_in_range(
    formula_cells: dict[str, str], row_start: int, row_end: int, col_start: int, col_end: int
) -> dict[str, str]:
    selected: dict[str, str] = {}
    for coordinate, formula in formula_cells.items():
        column_label = "".join(character for character in coordinate if character.isalpha())
        row_label = "".join(character for character in coordinate if character.isdigit())
        if not column_label or not row_label:
            continue
        row_number = int(row_label)
        column_number = _column_number(column_label)
        if row_start <= row_number <= row_end and col_start <= column_number <= col_end:
            selected[coordinate] = formula
    return selected


def _markdown_cell(value: Any) -> str:
    text = _safe_scalar(value).replace("\n", "<br>")
    return text.replace("|", "\\|")


def _chunk_limit_chars(params: TextChunkParams | None) -> int:
    params = params or TextChunkParams()
    # Existing text chunk params are token based.  A conservative character
    # estimate keeps tables row-aligned without requiring a tokenizer here.
    return max(512, int(params.max_tokens) * 4)


def _row_groups(rows: Sequence[int], values: dict[tuple[int, int], str], col_start: int, col_end: int, limit: int) -> list[list[int]]:
    groups: list[list[int]] = []
    current: list[int] = []
    current_size = 0
    for row_number in rows:
        row_size = sum(len(values.get((row_number, col), "")) + 3 for col in range(col_start, col_end + 1))
        if current and current_size + row_size > limit:
            groups.append(current)
            current = []
            current_size = 0
        current.append(row_number)
        current_size += row_size
    if current:
        groups.append(current)
    return groups


def _table_markdown(
    values: dict[tuple[int, int], str],
    rows: Sequence[int],
    col_start: int,
    col_end: int,
) -> str:
    """Render a rectangular region while preserving empty/ragged cells."""
    if not rows:
        return ""
    columns = list(range(col_start, col_end + 1))
    first = rows[0]
    header = [values.get((first, col), "") or f"Column {_column_name(col)}" for col in columns]
    lines = ["| " + " | ".join(_markdown_cell(value) for value in header) + " |"]
    lines.append("| " + " | ".join("---" for _ in columns) + " |")
    for row_number in rows[1:]:
        lines.append(
            "| " + " | ".join(_markdown_cell(values.get((row_number, col), "")) for col in columns) + " |"
        )
    return "\n".join(lines)


def _regions_from_values(values: dict[tuple[int, int], str]) -> list[tuple[list[int], int, int]]:
    """Split a sheet into row bands and side-by-side non-empty column bands."""
    by_row: dict[int, list[int]] = defaultdict(list)
    for row, col in values:
        by_row[row].append(col)
    sorted_rows = sorted(by_row)
    row_bands: list[list[int]] = []
    for row in sorted_rows:
        if not row_bands or row - row_bands[-1][-1] > 1:
            row_bands.append([row])
        else:
            row_bands[-1].append(row)

    regions: list[tuple[list[int], int, int]] = []
    for rows in row_bands:
        columns = sorted({col for row in rows for col in by_row[row]})
        col_bands: list[list[int]] = []
        for col in columns:
            if not col_bands or col - col_bands[-1][-1] > 1:
                col_bands.append([col])
            else:
                col_bands[-1].append(col)
        for band in col_bands:
            regions.append((rows, band[0], band[-1]))
    return regions


def _object_anchor(anchor: Any) -> str | None:
    if isinstance(anchor, str):
        return anchor
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    if start is None:
        return None
    start_ref = f"{_column_name(int(start.col) + 1)}{int(start.row) + 1}"
    if end is None:
        return start_ref
    end_ref = f"{_column_name(int(end.col) + 1)}{int(end.row) + 1}"
    return f"{start_ref}:{end_ref}"


def _chart_title(chart: Any) -> str:
    try:
        paragraphs = chart.title.tx.rich.p
        values = []
        for paragraph in paragraphs:
            for run in getattr(paragraph, "r", ()):
                text = getattr(run, "t", None)
                if text:
                    values.append(str(text))
        if values:
            return " ".join(values).strip()
    except Exception:
        pass
    return chart.__class__.__name__


def _workbook_objects(worksheet: Any) -> list[dict[str, Any]]:
    objects: list[dict[str, Any]] = []
    for index, image in enumerate(getattr(worksheet, "_images", ())):
        payload = None
        try:
            payload = image._data()
        except Exception:
            pass
        objects.append(
            {
                "type": "image",
                "index": index,
                "anchor": _object_anchor(getattr(image, "anchor", None)),
                "format": getattr(image, "format", None),
                "size_bytes": len(payload) if payload else None,
                "sha256": hashlib.sha256(payload).hexdigest() if payload else None,
                "needs_ocr": True,
            }
        )
    for index, chart in enumerate(getattr(worksheet, "_charts", ())):
        objects.append(
            {
                "type": "chart",
                "index": index,
                "anchor": _object_anchor(getattr(chart, "anchor", None)),
                "title": _chart_title(chart),
                "needs_ocr": False,
            }
        )
    return objects


def _load_openpyxl_workbooks(payload: bytes, path: str) -> tuple[Any, Any, str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised in minimal images
        raise RuntimeError("XLSX ingest requires the 'openpyxl' package") from exc

    extension = Path(path).suffix.lower()
    converted_path: str | None = None
    workbook_payload = payload
    if extension == ".xls":
        if shutil.which("libreoffice") is None:
            raise RuntimeError("XLS ingest requires LibreOffice to convert legacy .xls to .xlsx")
        with tempfile.TemporaryDirectory(prefix="retriever_xls_") as tmp_dir:
            source = Path(tmp_dir) / "input.xls"
            source.write_bytes(payload)
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx", str(source), "--outdir", tmp_dir],
                check=True,
                capture_output=True,
                text=True,
            )
            converted = Path(tmp_dir) / "input.xlsx"
            if not converted.is_file():
                raise RuntimeError("LibreOffice produced no XLSX output for the .xls workbook")
            workbook_payload = converted.read_bytes()
    elif extension != ".xlsx":
        raise ValueError(f"Unsupported spreadsheet workbook extension: {extension!r}")

    formula_workbook = load_workbook(io.BytesIO(workbook_payload), data_only=False, read_only=False)
    values_workbook = load_workbook(io.BytesIO(workbook_payload), data_only=True, read_only=False)
    return formula_workbook, values_workbook, converted_path


def _xlsx_to_rows(payload: bytes, path: str, params: TextChunkParams | None) -> list[dict[str, Any]]:
    formula_workbook, values_workbook, _ = _load_openpyxl_workbooks(payload, path)
    rows: list[dict[str, Any]] = []
    sheet_inventory = [
        {"sheet_name": sheet.title, "sheet_index": index + 1, "state": sheet.sheet_state}
        for index, sheet in enumerate(formula_workbook.worksheets)
    ]
    source_path = str(Path(path).resolve()) if Path(path).exists() else str(path)
    limit = _chunk_limit_chars(params)
    try:
        for sheet_index, (worksheet, values_sheet) in enumerate(
            zip(formula_workbook.worksheets, values_workbook.worksheets, strict=False), start=1
        ):
            cells: dict[tuple[int, int], str] = {}
            formula_cells: dict[str, str] = {}
            for cell in getattr(worksheet, "_cells", {}).values():
                formula_value = cell.value
                if formula_value is None:
                    continue
                coordinate = (int(cell.row), int(cell.column))
                value_cell = values_sheet.cell(row=cell.row, column=cell.column)
                display_value = value_cell.value if isinstance(formula_value, str) and formula_value.startswith("=") else formula_value
                cells[coordinate] = _safe_scalar(display_value)
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formula_cells[f"{_column_name(cell.column)}{cell.row}"] = formula_value

            objects = _workbook_objects(worksheet)
            regions = _regions_from_values(cells)
            for block_index, (region_rows, col_start, col_end) in enumerate(regions):
                for chunk_index, row_group in enumerate(
                    _row_groups(region_rows, cells, col_start, col_end, limit)
                ):
                    if not row_group:
                        continue
                    start_row, end_row = row_group[0], row_group[-1]
                    cell_range = _cell_range(start_row, end_row, col_start, col_end)
                    markdown = _table_markdown(cells, row_group, col_start, col_end)
                    text = (
                        f"Workbook: {Path(path).name}\n"
                        f"Sheet: {worksheet.title}\n"
                        f"Range: {cell_range}\n\n{markdown}"
                    )
                    metadata = {
                        "source_path": source_path,
                        "page_number": sheet_index,
                        "content_metadata": {
                            "type": "spreadsheet",
                            "source_type": "native_cell",
                            "reader_backend": "openpyxl",
                            "sheet_name": worksheet.title,
                            "sheet_index": sheet_index,
                            "range": cell_range,
                            "row_start": start_row,
                            "row_end": end_row,
                            "column_start": col_start,
                            "column_end": col_end,
                            "chunk_index": chunk_index,
                            "block_index": block_index,
                            "formula_cells": _formula_cells_in_range(
                                formula_cells, start_row, end_row, col_start, col_end
                            ),
                            "embedded_objects": objects,
                            "sheet_inventory": sheet_inventory,
                        },
                    }
                    rows.append(
                        {
                            "text": text,
                            "content": text,
                            "path": source_path,
                            "source_path": source_path,
                            "page_number": sheet_index,
                            "sheet_name": worksheet.title,
                            "chunk_index": len(rows),
                            "metadata": metadata,
                        }
                    )

            # Keep charts/images visible to retrieval and debugging even when
            # the sheet has no native cells. Their binary assets are not put in
            # the text row; the metadata carries stable anchor/hash provenance.
            for object_index, obj in enumerate(objects):
                if not obj.get("anchor"):
                    continue
                object_text = f"Workbook: {Path(path).name}\nSheet: {worksheet.title}\n"
                if obj["type"] == "chart":
                    object_text += f"Chart at {obj['anchor']}: {obj.get('title') or 'chart'}"
                else:
                    object_text += f"Embedded image at {obj['anchor']} (OCR candidate)"
                metadata = {
                    "source_path": source_path,
                    "page_number": sheet_index,
                    "content_metadata": {
                        "type": "spreadsheet",
                        "source_type": "chart_data" if obj["type"] == "chart" else "embedded_image",
                        "reader_backend": "openpyxl",
                        "sheet_name": worksheet.title,
                        "sheet_index": sheet_index,
                        "object_index": object_index,
                        "object": obj,
                        "sheet_inventory": sheet_inventory,
                    },
                }
                rows.append(
                    {
                        "text": object_text,
                        "content": object_text,
                        "path": source_path,
                        "source_path": source_path,
                        "page_number": sheet_index,
                        "sheet_name": worksheet.title,
                        "chunk_index": len(rows),
                        "metadata": metadata,
                    }
                )
    finally:
        formula_workbook.close()
        values_workbook.close()
    return rows


def _decode_csv(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace"), "utf-8-replace"


def _csv_to_rows(payload: bytes, path: str, params: TextChunkParams | None) -> list[dict[str, Any]]:
    raw_text, encoding = _decode_csv(payload)
    delimiter = _detect_csv_delimiter(raw_text)
    records = [[_safe_scalar(value) for value in row] for row in csv.reader(io.StringIO(raw_text), delimiter=delimiter)]
    records = [row for row in records if any(value.strip() for value in row)]
    if not records:
        return []
    # The header defines the logical schema. A data URI can contain the
    # delimiter itself (for example ``data:image/png;base64,...``), so a
    # malformed/unquoted row must not silently create a new fake column.
    width = max(1, len(records[0]))
    header_values = records[0] + [""] * (width - len(records[0]))
    header = [value or f"Column {_column_name(index + 1)}" for index, value in enumerate(header_values[:width])]
    image_header_indexes = {
        index
        for index, value in enumerate(header)
        if any(token in value.lower() for token in ("image", "photo", "picture", "thumbnail", "img"))
    }

    def normalize_row(row: list[str]) -> list[str]:
        if len(row) <= width:
            return row + [""] * (width - len(row))
        # Prefer the image column as the overflow sink. This preserves an
        # unquoted data URI; for other ragged rows the last logical column is
        # retained as a lossless delimiter-joined value.
        target = min(image_header_indexes) if image_header_indexes else width - 1
        suffix_count = width - target - 1
        image_end = len(row) - suffix_count
        merged = row[:target] + [delimiter.join(row[target:image_end])] + row[image_end:]
        return (merged + [""] * width)[:width]

    data_rows = [normalize_row(row) for row in records[1:]]
    limit = _chunk_limit_chars(params)
    source_path = str(Path(path).resolve()) if Path(path).exists() else str(path)
    image_columns = [header[index] for index in sorted(image_header_indexes)]
    rows: list[dict[str, Any]] = []
    current: list[list[str]] = []
    current_size = 0
    start_line = 2
    for offset, data_row in enumerate(data_rows):
        row_size = sum(len(value) + 3 for value in data_row)
        if current and current_size + row_size > limit:
            end_line = start_line + len(current) - 1
            range_name = f"row {start_line}:{end_line}"
            table = _table_markdown(
                {(1, index + 1): header[index] for index in range(width)}
                | {(index + 2, col + 1): value for index, data in enumerate(current) for col, value in enumerate(data)},
                [1] + list(range(2, len(current) + 2)),
                1,
                width,
            )
            rows.append(_csv_row(path, source_path, encoding, delimiter, range_name, table, image_columns, len(rows)))
            start_line = offset + 2
            current = []
            current_size = 0
        current.append(data_row)
        current_size += row_size
    if current:
        end_line = start_line + len(current) - 1
        range_name = f"row {start_line}:{end_line}"
        table = _table_markdown(
            {(1, index + 1): header[index] for index in range(width)}
            | {(index + 2, col + 1): value for index, data in enumerate(current) for col, value in enumerate(data)},
            [1] + list(range(2, len(current) + 2)),
            1,
            width,
        )
        rows.append(_csv_row(path, source_path, encoding, delimiter, range_name, table, image_columns, len(rows)))
    return rows


def _detect_csv_delimiter(raw_text: str) -> str:
    """Choose the delimiter with the most stable non-trivial row width.

    ``csv.Sniffer`` commonly mistakes the comma in a data URI for the CSV
    delimiter.  Scoring parsed row widths across several lines avoids that
    failure while still supporting semicolon/tab/pipe exports.
    """
    sample = "\n".join(line for line in raw_text.splitlines() if line.strip())[:16384]
    candidates = ",;\t|"
    best: tuple[int, int, int, str] | None = None
    for delimiter in candidates:
        try:
            parsed = list(csv.reader(io.StringIO(sample), delimiter=delimiter))
        except csv.Error:
            continue
        widths = [len(row) for row in parsed if row]
        if not widths:
            continue
        counts: dict[int, int] = defaultdict(int)
        for width in widths:
            counts[width] += 1
        stable_width, stable_lines = max(counts.items(), key=lambda item: (item[1], item[0]))
        if stable_width <= 1:
            continue
        # Prefer more consistently structured lines, then wider rows. The
        # delimiter itself is the final deterministic tie-breaker.
        score = (stable_lines, stable_width, -len(candidates) + candidates.index(delimiter), delimiter)
        if best is None or score > best:
            best = score
    return best[3] if best is not None else ","


def _csv_row(
    path: str,
    source_path: str,
    encoding: str,
    delimiter: str,
    range_name: str,
    table: str,
    image_columns: list[str],
    chunk_index: int,
) -> dict[str, Any]:
    text = f"CSV: {Path(path).name}\nRange: {range_name}\n\n{table}"
    metadata = {
        "source_path": source_path,
        "page_number": 1,
        "content_metadata": {
            "type": "spreadsheet",
            "source_type": "native_csv",
            "reader_backend": "python_csv",
            "sheet_name": "CSV",
            "range": range_name,
            "encoding": encoding,
            "delimiter": delimiter,
            "image_columns": image_columns,
            "chunk_index": chunk_index,
        },
    }
    return {
        "text": text,
        "content": text,
        "path": source_path,
        "source_path": source_path,
        "page_number": 1,
        "sheet_name": "CSV",
        "chunk_index": chunk_index,
        "metadata": metadata,
    }


def spreadsheet_bytes_to_chunks_df(
    content_bytes: bytes,
    path: str,
    params: TextChunkParams | None = None,
) -> pd.DataFrame:
    """Parse one XLSX/XLS/CSV payload into canonical text rows."""
    extension = Path(path).suffix.lower()
    if extension == ".csv":
        rows = _csv_to_rows(content_bytes, path, params)
    elif extension in WORKBOOK_EXTENSIONS:
        rows = _xlsx_to_rows(content_bytes, path, params)
    else:
        raise ValueError(f"Unsupported spreadsheet extension: {extension!r}")
    return pd.DataFrame(rows)
